#!/usr/bin/env python3
"""
measurement_ocr.py -- Extract numeric readings from a display filmed with a
handheld camera (multimeter, scale, gauge, thermometer, etc).

Cross-platform (Windows / macOS / Linux). Pure OpenCV GUI, no extra toolkit.

HOW IT WORKS
------------
1.  Open a video file (or a live camera index).
2.  On the first frame you click the 4 corners of the display (or press 'a'
    to let OpenCV try to auto-detect the largest quadrilateral).
3.  From then on, every frame is registered against that one fixed
    *reference* frame using ORB features + a RANSAC homography. Because
    every frame is matched to the SAME reference (not frame-to-frame),
    there is no drift -- this single homography both cancels hand-shake
    and de-skews the display into a flat, front-on rectangle, regardless
    of how the camera moves, as long as enough of the reference scene
    stays visible. A light exponential smoothing pass further damps
    residual jitter. If match quality against that reference degrades for
    a while (slow long-term drift) or is lost and then recovered, the
    tracker quietly adopts a fresh reference frame so it keeps working
    over long clips instead of freezing on a stale pose.
4.  Before OCR, recent rectified frames are triaged for blur (variance of
    Laplacian): either the sharpest one in a short rolling window is used,
    or several are median-stacked together (they're already pixel-aligned,
    since they all share the same rectified coordinate space) to fight
    noise and motion blur. The result is fed to an OCR backend (Tesseract
    by default, EasyOCR optionally with GPU) periodically, and the
    recognized text is drawn live over the video.
5.  Press SPACE to pause. Click (or click-drag) on the "Rectified" window
    to mark a numeric field -- useful when a display shows several values
    at once (e.g. voltage + current). Press 'e' to log the current
    reading(s) of all fields, with a timestamp, to a CSV file.
6.  A row of clickable buttons under the rectified view (Pause/Play, Clear,
    Log, Re-pick, OCR-/OCR+, Quit) mirrors every keyboard shortcut, so
    nothing requires the keyboard. A stats HUD in the corner of the
    "Original" window shows live FPS, match/inlier counts, rebase count,
    current frame sharpness, and field count; a green banner briefly
    confirms each "Log" action on screen (not just in the terminal).

KEYBINDINGS (main window)
--------------------------
  space         pause / resume
  click / drag  (in the "Rectified" window) define a new field
  n             rename the last-drawn field (typed in the terminal)
  c             clear all fields
  e             extract/log current field values to CSV
  r             re-pick the 4 corners (reset the tracking reference)
  +  /  -       run OCR more / less often while playing
  q  /  Esc     quit

Every one of these (except renaming, which needs typed input) also has a
matching on-screen button in the toolbar strip under the "Rectified" view --
click Pause/Play, Clear, Log, Re-pick, OCR-/OCR+, or Quit directly, no
keyboard needed.

The "Original" window's status line shows "[reference rebased xN]" whenever
the tracker has quietly adopted a fresh reference frame to keep up with
slow drift -- that's diagnostic, not something you need to act on.

INSTALL
-------
  pip install opencv-python numpy pytesseract
  # plus the Tesseract OCR binary itself (pytesseract only wraps it):
  #   Windows : https://github.com/UB-Mannheim/tesseract/wiki
  #   macOS   : brew install tesseract
  #   Linux   : sudo apt install tesseract-ocr   (or your distro's package)
  #
  # Optional GPU / alternative backend (no external binary needed):
  #   pip install easyocr
  #   python measurement_ocr.py video.mp4 --ocr easyocr --gpu

USAGE
-----
  python measurement_ocr.py video.mp4
  python measurement_ocr.py 0                      # live camera 0
  python measurement_ocr.py video.mp4 --ocr easyocr --gpu
"""

import argparse
import csv
import os
import sys
import time
from collections import deque
from dataclasses import dataclass, field as dataclass_field
from pathlib import Path
from typing import List, Optional, Tuple

import cv2
import numpy as np


# --------------------------------------------------------------------------
# Geometry helpers
# --------------------------------------------------------------------------

def order_points(pts: np.ndarray) -> np.ndarray:
    """Order 4 points as top-left, top-right, bottom-right, bottom-left."""
    pts = np.asarray(pts, dtype=np.float32)
    s = pts.sum(axis=1)
    d = np.diff(pts, axis=1).ravel()
    ordered = np.zeros((4, 2), dtype=np.float32)
    ordered[0] = pts[np.argmin(s)]   # top-left     (smallest x+y)
    ordered[2] = pts[np.argmax(s)]   # bottom-right (largest x+y)
    ordered[1] = pts[np.argmin(d)]   # top-right    (smallest y-x)
    ordered[3] = pts[np.argmax(d)]   # bottom-left  (largest y-x)
    return ordered


CANONICAL_MAX = 6000  # hard cap so a mis-dragged / degenerate quad can't ask
                      # warpPerspective to allocate a giant rectified image


def canonical_size_from_quad(quad: np.ndarray) -> Tuple[int, int]:
    """Pick an output (width, height) for the rectified image based on the
    real proportions of the quad the user selected. Clamped to sane bounds so
    a degenerate or far-dragged selection can't produce a 0-size or absurdly
    huge canvas (the latter would crash warpPerspective on allocation)."""
    tl, tr, br, bl = quad
    w = int(max(np.linalg.norm(tr - tl), np.linalg.norm(br - bl)))
    h = int(max(np.linalg.norm(bl - tl), np.linalg.norm(br - tr)))
    return min(max(w, 60), CANONICAL_MAX), min(max(h, 40), CANONICAL_MAX)


def auto_detect_quad(frame: np.ndarray) -> Optional[np.ndarray]:
    """Best-effort automatic detection of the largest rectangular panel in
    the frame (e.g. a device bezel). Falls back to None if nothing found;
    the caller should ask the user to click manually in that case."""
    gray = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)
    gray = cv2.GaussianBlur(gray, (5, 5), 0)
    edges = cv2.Canny(gray, 40, 120)
    edges = cv2.dilate(edges, np.ones((3, 3), np.uint8), iterations=1)
    contours, _ = cv2.findContours(edges, cv2.RETR_LIST, cv2.CHAIN_APPROX_SIMPLE)
    frame_area = frame.shape[0] * frame.shape[1]
    best, best_area = None, 0.0
    for c in contours:
        peri = cv2.arcLength(c, True)
        approx = cv2.approxPolyDP(c, 0.02 * peri, True)
        if len(approx) == 4 and cv2.isContourConvex(approx):
            area = cv2.contourArea(approx)
            if area > best_area and area > 0.02 * frame_area:
                best_area, best = area, approx.reshape(4, 2)
    return order_points(best) if best is not None else None


def resize_max_width(frame: np.ndarray, max_width: int) -> np.ndarray:
    if max_width and frame.shape[1] > max_width:
        scale = max_width / frame.shape[1]
        frame = cv2.resize(frame, (max_width, int(round(frame.shape[0] * scale))),
                            interpolation=cv2.INTER_AREA)
    return frame


# --------------------------------------------------------------------------
# Reference-frame tracker: drift-free deskew + antishake in one homography
# --------------------------------------------------------------------------

class ReferenceTracker:
    """
    Registers every incoming frame against one fixed reference frame using
    ORB features + a RANSAC homography, so the mapping from "current frame"
    to "flattened display" doesn't drift frame-to-frame, no matter how much
    the handheld camera shakes or rotates -- as long as enough of the
    reference scene stays visible. A short exponential smoothing pass on
    the homography further damps residual per-frame jitter.

    To tolerate slow, long-term drift (the view gradually moving far enough
    from the original reference that ORB matches thin out), the tracker can
    quietly *rebase*: adopt a recent well-tracked frame as the new
    reference and recompute its features, while keeping the canonical
    output coordinates perfectly continuous (a rebase changes which image
    is used for matching, never what "canonical" means). This happens
    either proactively, after several consecutive weakly-matched frames, or
    reactively, right after recovering from a tracking loss.
    """

    HARD_FAIL_MATCHES = 12     # below this many good matches -> tracking lost this frame
    MIN_INLIER_RATIO = 0.2     # inliers/good-matches below this -> fit isn't trustworthy
    RECOVERY_STREAK = 5        # this many consecutive lost frames counts as "was lost"
    MIN_AREA_RATIO = 0.2       # a plausible quad can't shrink...
    MAX_AREA_RATIO = 5.0       # ...or grow past these ratios frame-to-frame
    # "Weak" match quality is judged relative to this scene's own recent best
    # inlier count, not a fixed number -- a scene that only ever achieves 40
    # inliers and one that achieves 900 both need to be flagged well before
    # they hit their respective failure point, and a fixed threshold can't
    # do that for both.
    WEAK_RATIO_OF_PEAK = 0.35
    PEAK_DECAY = 0.98

    def __init__(self, reference_frame: np.ndarray, quad: np.ndarray,
                 canonical_size: Tuple[int, int], n_features: int = 1500,
                 smoothing: float = 0.6, rebase_after_weak: int = 20,
                 rebase_on_recovery: bool = True):
        self.canonical_w, self.canonical_h = canonical_size
        self.smoothing = float(np.clip(smoothing, 0.0, 0.97))
        self.rebase_after_weak = int(rebase_after_weak) if rebase_after_weak else 0
        self.rebase_on_recovery = bool(rebase_on_recovery)

        self.orb = cv2.ORB_create(nfeatures=n_features)
        self.matcher = cv2.BFMatcher(cv2.NORM_HAMMING)

        self._dst_canonical_corners = np.array(
            [[0, 0], [self.canonical_w - 1, 0],
             [self.canonical_w - 1, self.canonical_h - 1],
             [0, self.canonical_h - 1]], dtype=np.float32)

        ref_to_canonical = cv2.getPerspectiveTransform(
            quad.astype(np.float32), self._dst_canonical_corners)
        if not self._set_reference(reference_frame, ref_to_canonical, quad.astype(np.float32)):
            raise RuntimeError(
                "Not enough texture/detail around the reference frame to "
                "track it reliably. Pick a sharper frame with more visible "
                "detail near the display and try again.")

        self.last_good_H = self.ref_to_canonical.copy()
        self.smoothed_H = self.ref_to_canonical.copy()
        self.last_good_quad_in_frame = self.ref_quad.copy()
        self.last_good_area = float(cv2.contourArea(self.ref_quad))
        self.lost_streak = 0
        self.consec_weak = 0
        self.rebase_count = 0
        self.last_good_matches = 0
        self.last_inlier_count = 0
        self.peak_inliers = 0.0

    def _quad_plausible(self, quad: np.ndarray) -> bool:
        """Reject geometrically nonsensical homographies (self-intersecting,
        or wildly different area from the last accepted frame) that a
        borderline RANSAC inlier count can occasionally still produce,
        especially at oblique viewing angles -- these are worse than an
        honest failure because they'd otherwise be accepted as "tracked"."""
        quad32 = quad.astype(np.float32)
        if not np.all(np.isfinite(quad32)):
            return False
        if not cv2.isContourConvex(quad32.reshape(-1, 1, 2).astype(np.int32)):
            return False
        area = cv2.contourArea(quad32)
        if area <= 1.0:
            return False
        ratio = area / max(self.last_good_area, 1.0)
        return self.MIN_AREA_RATIO <= ratio <= self.MAX_AREA_RATIO

    def _set_reference(self, reference_frame: np.ndarray, ref_to_canonical: np.ndarray,
                        ref_quad: np.ndarray) -> bool:
        ref_gray = cv2.cvtColor(reference_frame, cv2.COLOR_BGR2GRAY)
        kp, des = self.orb.detectAndCompute(ref_gray, None)
        if des is None or len(kp) < 20:
            return False
        self.ref_kp, self.ref_des = kp, des
        self.ref_to_canonical = ref_to_canonical
        self.ref_quad = ref_quad
        return True

    def _rebase(self, frame: np.ndarray, H_cur_to_canonical: np.ndarray) -> None:
        """Adopt `frame` as the new reference, in a way that keeps canonical
        coordinates identical to before -- only the ORB matching target
        changes, never the output space."""
        new_ref_to_canonical = H_cur_to_canonical.copy()
        try:
            inv = np.linalg.inv(new_ref_to_canonical)
        except np.linalg.LinAlgError:
            return
        new_ref_quad = cv2.perspectiveTransform(
            self._dst_canonical_corners.reshape(-1, 1, 2), inv).reshape(4, 2)
        if self._set_reference(frame, new_ref_to_canonical, new_ref_quad):
            self.rebase_count += 1
            self.consec_weak = 0
            self.peak_inliers = 0.0
            self.last_good_area = float(cv2.contourArea(new_ref_quad.astype(np.float32)))

    def register(self, frame: np.ndarray) -> Tuple[np.ndarray, np.ndarray, bool]:
        """Returns (H_current_to_canonical, quad_in_current_frame, tracked_ok).
        On a bad/blurred frame it silently reuses the last good result so
        the overlay doesn't flicker or jump."""
        gray = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)
        kp, des = self.orb.detectAndCompute(gray, None)
        if des is None or len(kp) < 10:
            self.lost_streak += 1
            return self.last_good_H, self.last_good_quad_in_frame, False

        raw_matches = self.matcher.knnMatch(self.ref_des, des, k=2)
        good = [m for m, n in raw_matches if m.distance < 0.75 * n.distance]
        self.last_good_matches = len(good)
        if len(good) < self.HARD_FAIL_MATCHES:
            self.lost_streak += 1
            return self.last_good_H, self.last_good_quad_in_frame, False

        ref_pts = np.float32([self.ref_kp[m.queryIdx].pt for m in good]).reshape(-1, 1, 2)
        cur_pts = np.float32([kp[m.trainIdx].pt for m in good]).reshape(-1, 1, 2)

        H_cur_to_ref, mask = cv2.findHomography(cur_pts, ref_pts, cv2.RANSAC, 4.0)
        if H_cur_to_ref is None or mask is None:
            self.lost_streak += 1
            return self.last_good_H, self.last_good_quad_in_frame, False

        inlier_count = int(mask.sum())
        self.last_inlier_count = inlier_count
        inlier_ratio = inlier_count / len(good)
        if inlier_count < 10 or inlier_ratio < self.MIN_INLIER_RATIO:
            self.lost_streak += 1
            return self.last_good_H, self.last_good_quad_in_frame, False

        try:
            H_ref_to_cur = np.linalg.inv(H_cur_to_ref)
            quad_in_frame_candidate = cv2.perspectiveTransform(
                self.ref_quad.reshape(-1, 1, 2), H_ref_to_cur).reshape(4, 2)
        except np.linalg.LinAlgError:
            self.lost_streak += 1
            return self.last_good_H, self.last_good_quad_in_frame, False

        if not self._quad_plausible(quad_in_frame_candidate):
            # A borderline-inlier-count fit occasionally still produces a
            # geometrically nonsensical result, especially at oblique
            # angles -- treat this exactly like a failed match, not a
            # successful (but wrong) one.
            self.lost_streak += 1
            return self.last_good_H, self.last_good_quad_in_frame, False

        # We tracked successfully this frame. Note whether we're *recovering*
        # from a loss before resetting the counter -- that changes how we
        # blend/smooth below and whether we rebase.
        recovered_from_loss = self.lost_streak >= self.RECOVERY_STREAK
        was_tracking_cleanly = self.lost_streak == 0

        H_cur_to_canonical = self.ref_to_canonical @ H_cur_to_ref
        H_cur_to_canonical = H_cur_to_canonical / H_cur_to_canonical[2, 2]

        if self.smoothing > 0 and was_tracking_cleanly:
            self.smoothed_H = (self.smoothing * self.smoothed_H
                               + (1 - self.smoothing) * H_cur_to_canonical)
            H_use = self.smoothed_H.copy()
        else:
            # Don't smooth across a recovery -- that would blend a stale
            # pre-loss pose into a fresh post-loss one.
            H_use = H_cur_to_canonical
            self.smoothed_H = H_use.copy()

        quad_in_frame = quad_in_frame_candidate
        self.last_good_H = H_use
        self.last_good_quad_in_frame = quad_in_frame
        self.last_good_area = float(cv2.contourArea(quad_in_frame.astype(np.float32)))

        is_weak = self.peak_inliers > 0 and inlier_count < self.peak_inliers * self.WEAK_RATIO_OF_PEAK
        self.consec_weak = self.consec_weak + 1 if is_weak else 0
        self.peak_inliers = max(inlier_count, self.peak_inliers * self.PEAK_DECAY)

        do_rebase = ((self.rebase_on_recovery and recovered_from_loss) or
                     (self.rebase_after_weak and self.consec_weak >= self.rebase_after_weak))

        self.lost_streak = 0

        if do_rebase:
            self._rebase(frame, H_use)

        return H_use, quad_in_frame, True


# --------------------------------------------------------------------------
# OCR preprocessing + backends
# --------------------------------------------------------------------------

def preprocess_for_ocr(image: np.ndarray, upscale: float = 3.0,
                        auto_invert: bool = True) -> np.ndarray:
    """Grayscale -> upscale -> contrast boost -> binarize -> polarity fix.
    Tuned for small, low-res digit displays (LCD/LED/e-ink)."""
    gray = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY) if image.ndim == 3 else image.copy()
    if upscale and upscale != 1.0:
        gray = cv2.resize(gray, None, fx=upscale, fy=upscale, interpolation=cv2.INTER_CUBIC)
    clahe = cv2.createCLAHE(clipLimit=2.5, tileGridSize=(8, 8))
    gray = clahe.apply(gray)
    gray = cv2.GaussianBlur(gray, (3, 3), 0)
    _, th = cv2.threshold(gray, 0, 255, cv2.THRESH_BINARY + cv2.THRESH_OTSU)
    if auto_invert:
        # Tesseract wants dark text on a light background; digits are
        # usually the minority of pixels, so use that to fix polarity.
        if np.count_nonzero(th == 255) < np.count_nonzero(th == 0):
            th = 255 - th
    th = cv2.morphologyEx(th, cv2.MORPH_CLOSE, np.ones((2, 2), np.uint8))
    return th


class OcrBackend:
    def recognize(self, image: np.ndarray):
        """Return a list of ((x, y, w, h), text, confidence 0-1)."""
        raise NotImplementedError


def _ensure_tesseract_cmd(pytesseract) -> None:
    """pytesseract calls the bare `tesseract` command, which fails if the
    binary is installed but not on PATH (common on Windows). If it's not
    resolvable, fall back to the usual install locations."""
    import os
    import shutil
    if shutil.which(pytesseract.pytesseract.tesseract_cmd):
        return
    for cand in (
        r"C:\Program Files\Tesseract-OCR\tesseract.exe",
        r"C:\Program Files (x86)\Tesseract-OCR\tesseract.exe",
        "/opt/homebrew/bin/tesseract",
        "/usr/local/bin/tesseract",
        "/usr/bin/tesseract",
    ):
        if os.path.exists(cand):
            pytesseract.pytesseract.tesseract_cmd = cand
            return


class TesseractBackend(OcrBackend):
    def __init__(self, whitelist: str = "0123456789.-:", psm: int = 11):
        try:
            import pytesseract
        except ImportError as e:
            raise RuntimeError(
                "pytesseract not installed. `pip install pytesseract`, and "
                "install the Tesseract binary for your OS (see the header "
                "of this script).") from e
        self.pytesseract = pytesseract
        _ensure_tesseract_cmd(pytesseract)
        # PSM 11 ("sparse text") finds numbers scattered across a multi-value
        # panel; PSM 7 ("single line") returns *nothing* on a multi-row
        # display, which is what produced empty CSVs. An empty whitelist lets
        # letters through too (e.g. to also read parameter-name labels).
        wl = f' -c tessedit_char_whitelist={whitelist}' if whitelist else ''
        self.config = f'--psm {psm}{wl}'

    def recognize(self, image):
        data = self.pytesseract.image_to_data(
            image, config=self.config, output_type=self.pytesseract.Output.DICT)
        results = []
        for i in range(len(data['text'])):
            text = data['text'][i].strip()
            if not text:
                continue
            try:
                conf = max(0.0, float(data['conf'][i])) / 100.0
            except (ValueError, TypeError):
                conf = 0.0
            box = (data['left'][i], data['top'][i], data['width'][i], data['height'][i])
            results.append((box, text, conf))
        return results


class EasyOcrBackend(OcrBackend):
    def __init__(self, gpu: bool = False, allowlist: str = "0123456789.-:"):
        try:
            import easyocr
        except ImportError as e:
            raise RuntimeError(
                "easyocr not installed. `pip install easyocr` (this pulls "
                "in PyTorch; use --gpu to run it on CUDA if available).") from e
        self.reader = easyocr.Reader(['en'], gpu=gpu, verbose=False)
        self.allowlist = allowlist

    def recognize(self, image):
        out = self.reader.readtext(image, allowlist=self.allowlist or None)
        results = []
        for box, text, conf in out:
            xs, ys = [p[0] for p in box], [p[1] for p in box]
            x, y = int(min(xs)), int(min(ys))
            w, h = int(max(xs) - x), int(max(ys) - y)
            text = text.strip()
            if text:
                results.append(((x, y, w, h), text, float(conf)))
        return results


def build_ocr_backend(args) -> OcrBackend:
    if args.ocr == "easyocr":
        return EasyOcrBackend(gpu=args.gpu, allowlist=args.whitelist)
    return TesseractBackend(whitelist=args.whitelist, psm=args.psm)


# --------------------------------------------------------------------------
# Blur handling: sharpness-gated frame selection + registered frame stacking
# --------------------------------------------------------------------------

def sharpness_score(image: np.ndarray) -> float:
    """Variance of the Laplacian -- a standard, cheap focus/blur metric.
    Higher is sharper; the absolute value only means something relative to
    other frames of the same scene, so it's used for ranking, not as a
    fixed universal threshold."""
    gray = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY) if image.ndim == 3 else image
    return float(cv2.Laplacian(gray, cv2.CV_64F).var())


def pick_best_frame(buf: deque, n: int) -> Optional[np.ndarray]:
    """Sharpest frame among the last n buffered (image, sharpness) pairs."""
    items = list(buf)[-n:]
    if not items:
        return None
    return max(items, key=lambda t: t[1])[0]


def stacked_frame(buf: deque, n: int, min_relative_sharpness: float = 0.35) -> Optional[np.ndarray]:
    """Median-stack the last n buffered frames. They're already pixel-aligned
    (all warped into the same canonical coordinate space), so no extra
    registration is needed. The median is naturally robust to a minority of
    motion-blurred/noisy frames in the window; frames far blurrier than the
    sharpest one in the window are additionally excluded outright so they
    can't drag the result down. This helps with blur/noise that *varies*
    frame to frame -- it can't sharpen a lens that's out of focus in every
    single frame, since that blur is identical (and so survives) in all of
    them."""
    items = list(buf)[-n:]
    if not items:
        return None
    max_sharp = max(s for _, s in items)
    floor = max_sharp * min_relative_sharpness
    imgs = [img for img, s in items if s >= floor] or [img for img, _ in items]
    stack = np.stack(imgs, axis=0).astype(np.float32)
    return np.median(stack, axis=0).astype(np.uint8)


# --------------------------------------------------------------------------
# Fields (user-defined numeric regions inside the rectified display)
# --------------------------------------------------------------------------

@dataclass
class Field:
    name: str
    rect: Tuple[int, int, int, int]     # x, y, w, h in canonical/rectified coords
    last_text: str = ""
    last_conf: float = 0.0
    # Bounding boxes of the recognized value pieces, in canonical/rectified
    # coords (mapped back from the upscaled OCR image). Drawn in red.
    last_boxes: List[Tuple[int, int, int, int]] = dataclass_field(default_factory=list)
    # The recognized text of each value piece, index-aligned with last_boxes.
    # A field may hold several (e.g. the whole-panel field over a table); these
    # get split into one-number-per-column at log time via field_columns().
    last_tokens: List[str] = dataclass_field(default_factory=list)


def run_ocr_and_update_fields(ocr_source: np.ndarray, fields: List[Field],
                               backend: OcrBackend, args) -> None:
    """If no fields were defined yet, auto-creates one covering the whole
    rectified display, so OCR runs out of the box with zero clicks.
    `ocr_source` is the blur/noise-optimized image (sharpest-of-window or
    median stack) -- same canonical coordinates as the live "Rectified"
    view, so field rects apply unchanged."""
    if not fields:
        fields.append(Field(name="display",
                             rect=(0, 0, ocr_source.shape[1], ocr_source.shape[0])))
    for f in fields:
        x, y, w, h = f.rect
        crop = ocr_source[y:y + h, x:x + w]
        if crop.size == 0:
            continue
        proc = preprocess_for_ocr(crop, upscale=args.upscale,
                                   auto_invert=not args.no_auto_invert)
        proc_bgr = cv2.cvtColor(proc, cv2.COLOR_GRAY2BGR)
        try:
            results = backend.recognize(proc_bgr)
        except Exception:
            results = []
        if results:
            results.sort(key=lambda r: r[0][0])  # left to right
            f.last_text = " ".join(r[1] for r in results)
            f.last_conf = float(np.mean([r[2] for r in results]))
            # Map each value box from upscaled-crop space back to canonical
            # coords: undo the OCR upscale, then offset by the field origin.
            up = args.upscale if args.upscale else 1.0
            f.last_boxes = [
                (int(x + bx / up), int(y + by / up), int(bw / up), int(bh / up))
                for (bx, by, bw, bh), _text, _conf in results
            ]
            f.last_tokens = [text for _box, text, _conf in results]


def parse_grid_spec(spec: str) -> Tuple[int, int]:
    """Parse a '<rows>x<cols>' grid spec (also accepts X, * or , as the
    separator) into (rows, cols). Raises ValueError on anything malformed."""
    for sep in ("x", "X", "*", ","):
        if sep in spec:
            a, b = spec.split(sep, 1)
            rows, cols = int(a), int(b)
            if rows < 1 or cols < 1:
                raise ValueError("grid rows and cols must both be >= 1")
            return rows, cols
    raise ValueError("grid must look like RxC, e.g. 3x4")


def build_grid_fields(rows: int, cols: int, width: int, height: int) -> List[Field]:
    """Evenly tile the canonical display with rows*cols fields named r1c1,
    r1c2, ... (row-major). Each becomes its own stable, labeled column; the
    user can drag any cell's corner to fit it and rename it with 'n'."""
    fields: List[Field] = []
    for r in range(rows):
        y0 = int(round(r * height / rows))
        y1 = int(round((r + 1) * height / rows))
        for c in range(cols):
            x0 = int(round(c * width / cols))
            x1 = int(round((c + 1) * width / cols))
            fields.append(Field(name=f"r{r + 1}c{c + 1}", rect=(x0, y0, x1 - x0, y1 - y0)))
    return fields


# --------------------------------------------------------------------------
# Drawing
# --------------------------------------------------------------------------

def draw_original_overlay(frame: np.ndarray, quad_in_frame: np.ndarray,
                           tracked_ok: bool, paused: bool, rebase_count: int = 0) -> None:
    color = (0, 255, 0) if tracked_ok else (0, 0, 255)
    cv2.polylines(frame, [quad_in_frame.astype(np.int32)], True, color, 2)
    if paused:
        status = "PAUSED"
    elif tracked_ok:
        status = "TRACKING"
    else:
        status = "TRACKING LOST (holding last good pose)"
    if rebase_count:
        status += f"  [reference rebased x{rebase_count}]"
    cv2.putText(frame, status, (10, 25), cv2.FONT_HERSHEY_SIMPLEX, 0.7, color, 2, cv2.LINE_AA)


# Overlay colors (BGR). Purple outlines the parameter regions; red outlines
# the recognized value boxes inside them.
PARAM_COLOR = (200, 0, 140)   # purple
VALUE_COLOR = (0, 0, 255)     # red


def draw_rectified_overlay(img: np.ndarray, fields: List[Field], mouse_state: dict) -> None:
    if mouse_state.get("dragging") and mouse_state.get("start") and mouse_state.get("cur"):
        p1 = tuple(int(v) for v in mouse_state["start"])
        p2 = tuple(int(v) for v in mouse_state["cur"])
        cv2.rectangle(img, p1, p2, (255, 255, 0), 1)
    for f in fields:
        x, y, w, h = f.rect
        # Purple box around the parameter region the user/OCR found...
        cv2.rectangle(img, (x, y), (x + w, y + h), PARAM_COLOR, 2)
        # ...small purple corner handles you can grab to resize the field...
        for (hx, hy) in ((x, y), (x + w, y), (x, y + h), (x + w, y + h)):
            cv2.rectangle(img, (hx - 3, hy - 3), (hx + 3, hy + 3), PARAM_COLOR, -1)
        # ...and a red box around each recognized value piece inside it.
        for (vx, vy, vw, vh) in f.last_boxes:
            cv2.rectangle(img, (vx, vy), (vx + vw, vy + vh), VALUE_COLOR, 1)
        label = f"{f.name}: {f.last_text}  ({f.last_conf:.2f})"
        cv2.putText(img, label, (x, max(15, y - 6)), cv2.FONT_HERSHEY_SIMPLEX,
                    0.5, PARAM_COLOR, 1, cv2.LINE_AA)


# --------------------------------------------------------------------------
# Rudimentary on-screen UI: clickable toolbar buttons + a stats HUD panel
# --------------------------------------------------------------------------

TOOLBAR_HEIGHT = 40

# (internal command id, button label). "cmd_pause"'s label is swapped for
# "Play"/"Pause" at draw time depending on current state.
BUTTON_DEFS = [
    ("cmd_pause", "Pause"),
    ("cmd_clear", "Clear"),
    ("cmd_log", "Log"),
    ("cmd_rebase", "Re-pick"),
    ("cmd_ocr_slower", "OCR-"),
    ("cmd_ocr_faster", "OCR+"),
    ("cmd_quit", "Quit"),
]


@dataclass
class Button:
    cmd: str
    label: str
    rect: Tuple[int, int, int, int]  # x, y, w, h -- toolbar-local pixel coords


def build_toolbar_buttons(width: int) -> List[Button]:
    n = len(BUTTON_DEFS)
    bw = max(1, width // n)
    buttons, x = [], 0
    for i, (cmd, label) in enumerate(BUTTON_DEFS):
        w = bw if i < n - 1 else max(bw, width - x)  # last button absorbs rounding
        buttons.append(Button(cmd=cmd, label=label, rect=(x, 0, w, TOOLBAR_HEIGHT)))
        x += w
    return buttons


def hit_test_button(buttons: List[Button], x: int, y: int) -> Optional[str]:
    for b in buttons:
        bx, by, bw, bh = b.rect
        if bx <= x < bx + bw and by <= y < by + bh:
            return b.cmd
    return None


def draw_toolbar(width: int, buttons: List[Button], paused: bool,
                  pressed_cmd: Optional[str]) -> np.ndarray:
    """A row of clickable buttons rendered as a strip; composited below the
    rectified display. `pressed_cmd` briefly highlights whichever button was
    just clicked, as visual click feedback."""
    bar = np.full((TOOLBAR_HEIGHT, width, 3), (45, 45, 45), dtype=np.uint8)
    font = cv2.FONT_HERSHEY_SIMPLEX
    for b in buttons:
        x, y, w, h = b.rect
        label = ("Play" if paused else "Pause") if b.cmd == "cmd_pause" else b.label
        bg = (0, 145, 220) if b.cmd == pressed_cmd else (85, 85, 85)
        cv2.rectangle(bar, (x + 2, y + 3), (x + w - 3, y + h - 4), bg, -1)
        cv2.rectangle(bar, (x + 2, y + 3), (x + w - 3, y + h - 4), (15, 15, 15), 1)
        ts = cv2.getTextSize(label, font, 0.45, 1)[0]
        tx, ty = x + max(2, (w - ts[0]) // 2), y + (h + ts[1]) // 2
        cv2.putText(bar, label, (tx, ty), font, 0.45, (255, 255, 255), 1, cv2.LINE_AA)
    return bar


def compute_display_params(canonical_w: int, canonical_h: int,
                            target_width: int = 480, max_scale: float = 3.0) -> Tuple[float, int, int]:
    """The canonical rectified image is often small (a tightly cropped
    display); this picks an upscale factor so the on-screen window has
    enough room for legible buttons, without blowing up a huge source past
    a reasonable zoom."""
    scale = max(1.0, min(max_scale, target_width / max(canonical_w, 1)))
    disp_w = max(1, int(round(canonical_w * scale)))
    disp_h = max(1, int(round(canonical_h * scale)))
    return scale, disp_w, disp_h


def draw_stats_panel(frame: np.ndarray, lines: List[str],
                      corner: str = "top_right", margin: int = 10) -> None:
    """Semi-transparent HUD box listing live tracking/OCR stats."""
    if not lines:
        return
    pad, line_h = 8, 18
    font, font_scale = cv2.FONT_HERSHEY_SIMPLEX, 0.45
    box_w = max(cv2.getTextSize(l, font, font_scale, 1)[0][0] for l in lines) + pad * 2
    box_h = line_h * len(lines) + pad * 2
    h, w = frame.shape[:2]
    x0, y0 = (w - box_w - margin, margin) if corner == "top_right" else (margin, margin)
    overlay = frame.copy()
    cv2.rectangle(overlay, (x0, y0), (x0 + box_w, y0 + box_h), (20, 20, 20), -1)
    cv2.addWeighted(overlay, 0.6, frame, 0.4, 0, frame)
    for i, line in enumerate(lines):
        ty = y0 + pad + line_h * (i + 1) - 5
        cv2.putText(frame, line, (x0 + pad, ty), font, font_scale, (0, 255, 180), 1, cv2.LINE_AA)


def draw_flash_banner(frame: np.ndarray, text: str) -> None:
    """Brief, prominent on-screen confirmation (e.g. after logging), since a
    person watching the video window won't see terminal output."""
    font, font_scale = cv2.FONT_HERSHEY_SIMPLEX, 0.6
    (tw, th), _ = cv2.getTextSize(text, font, font_scale, 2)
    x, y, pad = max(10, (frame.shape[1] - tw) // 2), 40, 10
    overlay = frame.copy()
    cv2.rectangle(overlay, (x - pad, y - th - pad), (x + tw + pad, y + pad), (0, 120, 0), -1)
    cv2.addWeighted(overlay, 0.75, frame, 0.25, 0, frame)
    cv2.putText(frame, text, (x, y), font, font_scale, (255, 255, 255), 2, cv2.LINE_AA)


# --------------------------------------------------------------------------
# Interactive corner selection
# --------------------------------------------------------------------------

def select_corners_interactively(frame: np.ndarray, window: str = "Original") -> Optional[np.ndarray]:
    print("Click to drop the 4 display corners, then drag any corner to fine-tune. "
          "Press Enter to confirm, 'a' to auto-detect, 'q' to cancel.")
    pts: List[List[int]] = []
    drag = {"idx": None}
    handle_r = 8

    def nearest_corner(x, y):
        best, best_d = None, (handle_r * 2) ** 2
        for i, (px, py) in enumerate(pts):
            d = (px - x) ** 2 + (py - y) ** 2
            if d <= best_d:
                best, best_d = i, d
        return best

    def on_mouse(event, x, y, flags, userdata):
        if event == cv2.EVENT_LBUTTONDOWN:
            i = nearest_corner(x, y)
            if i is not None:
                drag["idx"] = i                 # grab an existing corner to move it
            elif len(pts) < 4:
                pts.append([x, y])              # or drop the next new corner
                drag["idx"] = len(pts) - 1
        elif event == cv2.EVENT_MOUSEMOVE and drag["idx"] is not None:
            pts[drag["idx"]] = [x, y]
        elif event == cv2.EVENT_LBUTTONUP:
            drag["idx"] = None

    cv2.setMouseCallback(window, on_mouse)
    while True:
        display = frame.copy()
        if len(pts) >= 2:
            cv2.polylines(display, [np.array(pts, dtype=np.int32)],
                          len(pts) == 4, (0, 255, 0), 2)
        for i, p in enumerate(pts):
            cv2.circle(display, tuple(p), handle_r, (0, 255, 0), 2)
            cv2.circle(display, tuple(p), 2, (0, 255, 0), -1)
            cv2.putText(display, str(i + 1), (p[0] + 10, p[1] - 10),
                        cv2.FONT_HERSHEY_SIMPLEX, 0.5, (0, 255, 0), 1, cv2.LINE_AA)
        hint = ("Enter=confirm  drag corners to adjust  a=auto  q=cancel"
                if len(pts) == 4 else f"Click corner {len(pts) + 1} of 4  (a=auto, q=cancel)")
        cv2.putText(display, hint, (10, 25), cv2.FONT_HERSHEY_SIMPLEX,
                    0.6, (0, 255, 255), 2, cv2.LINE_AA)
        cv2.imshow(window, display)
        key = cv2.waitKey(20) & 0xFF

        if key in (13, 10) and len(pts) == 4:   # Enter / Return confirms
            return order_points(np.array(pts, dtype=np.float32))
        if key == ord('a'):
            quad = auto_detect_quad(frame)
            if quad is not None:
                return quad
            print("Auto-detect couldn't find a clear rectangle -- click the corners manually.")
        if key in (ord('q'), 27):
            return None


def _grab_field_corner(fields: List[Field], cx: float, cy: float, tol: float):
    """If (cx, cy) is within `tol` of any field's corner, return
    (field, anchor_x, anchor_y) where the anchor is the diagonally-opposite
    corner (held fixed while the grabbed corner is dragged). Nearest wins."""
    best = None
    best_d = tol * tol
    for f in fields:
        x, y, w, h = f.rect
        # (grabbed corner) -> (opposite/anchor corner)
        for gx, gy, ax, ay in ((x, y, x + w, y + h),
                               (x + w, y, x, y + h),
                               (x, y + h, x + w, y),
                               (x + w, y + h, x, y)):
            d = (gx - cx) ** 2 + (gy - cy) ** 2
            if d <= best_d:
                best_d = d
                best = (f, ax, ay)
    return best


def make_composite_mouse_callback(state: dict, scale_ref: List[float], image_h_ref: List[int],
                                   buttons_ref: List[List[Button]],
                                   fields_ref: List[List[Field]], canon_ref: List[int]):
    """One callback serves the whole "Rectified" window: clicks below
    `image_h_ref[0]` (the toolbar strip) are hit-tested against the current
    button layout; clicks above it are field gestures, converted back from
    on-screen display pixels to canonical coordinates via `scale_ref[0]` so
    field rects stay correct regardless of display zoom. Grabbing near an
    existing field's corner *edits* that field (drag to resize); a drag on
    empty space *creates* a new field. Cell-style (single-element list) refs
    let the layout change (e.g. after a re-pick) without re-registering."""
    def _clamp(v, hi):
        return max(0.0, min(float(hi), v))

    def _end_gesture():
        state["dragging"] = False
        state["edit_field"] = None
        state["edit_anchor"] = None

    def on_mouse(event, x, y, flags, userdata):
        image_h = image_h_ref[0]
        if y >= image_h:
            if event == cv2.EVENT_LBUTTONDOWN:
                cmd = hit_test_button(buttons_ref[0], x, y - image_h)
                if cmd:
                    state["button_cmd"] = cmd
                    state["flash_cmd"] = cmd
                    state["flash_button_until"] = time.time() + 0.25
            elif event == cv2.EVENT_LBUTTONUP:
                _end_gesture()   # a drag that ends over the toolbar still ends cleanly
            return
        scale = scale_ref[0] or 1.0
        cx, cy = x / scale, y / scale
        cw, ch = canon_ref[0], canon_ref[1]

        if event == cv2.EVENT_LBUTTONDOWN:
            grab = _grab_field_corner(fields_ref[0], cx, cy, tol=max(6.0, 12.0 / scale))
            if grab is not None:
                f, ax, ay = grab
                state["edit_field"] = f
                state["edit_anchor"] = (ax, ay)
            else:
                state["dragging"] = True
                state["start"] = (cx, cy)
                state["cur"] = (cx, cy)
        elif event == cv2.EVENT_MOUSEMOVE:
            if state.get("edit_field") is not None:
                ax, ay = state["edit_anchor"]
                nx, ny = _clamp(cx, cw), _clamp(cy, ch)
                f = state["edit_field"]
                f.rect = (int(min(ax, nx)), int(min(ay, ny)),
                          int(abs(nx - ax)), int(abs(ny - ay)))
            elif state.get("dragging"):
                state["cur"] = (cx, cy)
        elif event == cv2.EVENT_LBUTTONUP:
            if state.get("edit_field") is not None:
                state["edit_field"] = None
                state["edit_anchor"] = None
            elif state.get("dragging"):
                state["dragging"] = False
                state["completed_rect"] = (state["start"][0], state["start"][1], cx, cy)
    return on_mouse


# --------------------------------------------------------------------------
# CSV logging
# --------------------------------------------------------------------------

def _row_major_tokens(tokens: List[str],
                      boxes: List[Tuple[int, int, int, int]]) -> List[str]:
    """Order (token, box) pairs into natural reading order -- top-to-bottom by
    row, left-to-right within a row -- so a table reads the way a person would.
    Falls back to the input order if boxes are missing or don't line up."""
    if not boxes or len(boxes) != len(tokens):
        return list(tokens)
    heights = sorted(h for (_x, _y, _w, h) in boxes)
    med_h = heights[len(heights) // 2] or 1
    row_tol = max(1.0, med_h * 0.6)
    order = sorted(range(len(tokens)),
                   key=lambda i: (round(boxes[i][1] / row_tol), boxes[i][0]))
    return [tokens[i] for i in order]


def field_columns(fields: List[Field]) -> Tuple[List[str], List[str]]:
    """Flatten fields into parallel (names, values) lists with ONE number per
    column. A field that recognized several numbers (e.g. the whole-panel field
    over a multi-value display) is split into name_1, name_2, ... in reading
    order; a single-value field stays one column under its own name."""
    names: List[str] = []
    values: List[str] = []
    for f in fields:
        toks = _row_major_tokens(f.last_tokens, f.last_boxes)
        if len(toks) <= 1:
            names.append(f.name)
            values.append(toks[0] if toks else f.last_text)
        else:
            for i, t in enumerate(toks, 1):
                names.append(f"{f.name}_{i}")
                values.append(t)
    return names, values


def _as_number(cell):
    """Coerce numeric-looking strings to real numbers so Excel treats them as
    values, not text. Non-numeric strings pass through; blanks become empty."""
    if cell is None or cell == "":
        return None
    if isinstance(cell, (int, float)):
        return cell
    s = str(cell)
    try:
        f = float(s)
    except ValueError:
        return s
    return int(f) if f.is_integer() and "." not in s and "e" not in s.lower() else f


class RowLogger:
    """Append-only tabular sink; CSV and XLSX share this interface so the
    logging code above doesn't care which format is in use."""
    def existing_header(self) -> Optional[List]:
        return None

    def append(self, cells: List) -> None:
        raise NotImplementedError

    def close(self) -> None:
        pass


class CsvLogger(RowLogger):
    def __init__(self, path: str):
        self._path = path
        self._file = open(path, "a", newline="")
        self._writer = csv.writer(self._file)

    def existing_header(self):
        if not (os.path.exists(self._path) and os.path.getsize(self._path) > 0):
            return None
        with open(self._path, newline="") as fh:
            return next(csv.reader(fh), None)

    def append(self, cells):
        self._writer.writerow(cells)
        self._file.flush()

    def close(self):
        self._file.close()


class XlsxLogger(RowLogger):
    def __init__(self, path: str):
        import openpyxl
        self._path = path
        self._had_data = os.path.exists(path) and os.path.getsize(path) > 0
        if self._had_data:
            self._wb = openpyxl.load_workbook(path)
            self._ws = self._wb.active
        else:
            self._wb = openpyxl.Workbook()
            self._ws = self._wb.active
            self._ws.title = "readings"
        self._save()

    def existing_header(self):
        # Only read row 1 when the file actually had data -- touching self._ws[1]
        # on a fresh sheet would materialize an empty row and shift the header.
        if not self._had_data or self._ws.max_row < 1:
            return None
        row = [c.value for c in self._ws[1]]
        return row if any(v is not None for v in row) else None

    def append(self, cells):
        self._ws.append([_as_number(c) for c in cells])
        self._save()

    def _save(self):
        self._wb.save(self._path)

    def close(self):
        self._save()


def build_row_logger(path: str) -> RowLogger:
    """Pick a logger from the output extension: .xlsx -> Excel, else CSV."""
    if str(path).lower().endswith((".xlsx", ".xlsm")):
        try:
            import openpyxl  # noqa: F401
        except ImportError as e:
            raise RuntimeError(
                "Writing .xlsx needs openpyxl. `pip install openpyxl`, "
                "or use a .csv --output path instead.") from e
        return XlsxLogger(str(path))
    return CsvLogger(str(path))


def log_row(logger: RowLogger, fields: List[Field], cap, frame_idx: int,
            log_state: dict) -> None:
    if not fields:
        print("No fields defined yet -- nothing to log.")
        return
    names, values = field_columns(fields)

    # Lock the column layout on the first row (or adopt the header of a file
    # we're appending to) so every later row lines up under the same columns.
    columns = log_state.get("columns")
    if columns is None:
        existing = logger.existing_header()
        if existing:
            columns = [str(c) for c in existing[2:]]  # drop timestamp_s, frame_idx
            log_state["header_written"] = True
        else:
            columns = names
        log_state["columns"] = columns
    # Align this row to the locked columns by name (missing -> blank, extras
    # dropped) so a frame that reads a different count doesn't shift the table.
    by_name: dict = {}
    for n, v in zip(names, values):
        by_name.setdefault(n, v)
    row_values = [by_name.get(col, "") for col in columns]

    ts = cap.get(cv2.CAP_PROP_POS_MSEC) / 1000.0
    try:
        if not log_state.get("header_written"):
            logger.append(["timestamp_s", "frame_idx"] + columns)
            log_state["header_written"] = True
        logger.append([f"{ts:.3f}", frame_idx] + row_values)
    except OSError as e:
        # e.g. the user opened the output in Excel mid-run and it's now locked.
        # Drop the row rather than crash; warn once so it isn't silent.
        if not log_state.get("warned_write"):
            print(f"Warning: could not write to the output file ({e}). "
                  "Is it open in another program? Rows are being dropped until "
                  "it's closed.")
            log_state["warned_write"] = True
        return
    log_state["warned_write"] = False
    # Dedupe key is the raw expansion, so auto-log fires only on real changes.
    log_state["last_values"] = values
    log_state["last_names"] = names
    log_state["rows_logged"] = log_state.get("rows_logged", 0) + 1
    print("Logged:", dict(zip(columns, row_values)))


def maybe_auto_log(ctx: dict) -> None:
    """Write a row automatically when the set of readings has changed since the
    last logged row (manual or automatic). Skips blank readings and unchanged
    ones so a played-through clip yields one row per distinct reading rather
    than a row per frame or an empty file."""
    fields = ctx['fields']
    if not fields:
        return
    names, values = field_columns(fields)
    if all(v is None or str(v).strip() == "" for v in values):
        return  # nothing recognized yet -- don't log a blank leading row
    log_state = ctx['log_state']
    # Re-log if the readings changed OR the column layout changed.
    if values == log_state.get("last_values") and names == log_state.get("last_names"):
        return
    log_row(ctx['logger'], fields, ctx['cap'], ctx['frame_idx'], log_state)


# --------------------------------------------------------------------------
# Action dispatch: keyboard shortcuts and toolbar button clicks both funnel
# through here, so the two input paths can never do subtly different things.
# --------------------------------------------------------------------------

COMMAND_KEYS = {
    ord(' '): 'toggle_pause',
    ord('c'): 'clear_fields',
    ord('e'): 'log',
    ord('r'): 'rebase',
    ord('+'): 'ocr_faster',
    ord('='): 'ocr_faster',   # unshifted key on the same physical key as '+'
    ord('-'): 'ocr_slower',
    ord('q'): 'quit',
    27: 'quit',
}

BUTTON_TO_ACTION = {
    'cmd_pause': 'toggle_pause',
    'cmd_clear': 'clear_fields',
    'cmd_log': 'log',
    'cmd_rebase': 'rebase',
    'cmd_ocr_slower': 'ocr_slower',
    'cmd_ocr_faster': 'ocr_faster',
    'cmd_quit': 'quit',
}


def refresh_display_params(ctx: dict) -> None:
    """Recompute the display scale/toolbar layout for the tracker's current
    canonical size -- called at startup and after any re-pick, since the
    canonical size (and therefore the toolbar width) can change."""
    scale, disp_w, disp_h = compute_display_params(ctx['tracker'].canonical_w, ctx['tracker'].canonical_h)
    ctx['scale_ref'][0] = scale
    ctx['image_h_ref'][0] = disp_h
    ctx['buttons_ref'][0] = build_toolbar_buttons(disp_w)
    ctx['canon_ref'][0] = ctx['tracker'].canonical_w
    ctx['canon_ref'][1] = ctx['tracker'].canonical_h
    ctx['disp_w'], ctx['disp_h'] = disp_w, disp_h


def apply_action(action: str, ctx: dict) -> bool:
    """Executes one command against the shared app state `ctx`. Returns
    True if the app should quit."""
    if action == 'toggle_pause':
        ctx['paused'] = not ctx['paused']
        ctx['dirty'] = True
    elif action == 'clear_fields':
        ctx['fields'].clear()
        ctx['dirty'] = True
    elif action == 'log':
        had_fields = bool(ctx['fields'])
        log_row(ctx['logger'], ctx['fields'], ctx['cap'],
                 ctx['frame_idx'], ctx['log_state'])
        if had_fields:
            names, values = field_columns(ctx['fields'])
            summary = ", ".join(f"{n}={v or '?'}" for n, v in zip(names, values))
            ctx['flash_text'] = f"Logged: {summary}"
            ctx['flash_until'] = time.time() + 1.5
    elif action == 'rebase':
        new_quad = select_corners_interactively(ctx['frame'])
        if new_quad is not None:
            canonical_size = canonical_size_from_quad(new_quad)
            try:
                ctx['tracker'] = ReferenceTracker(
                    ctx['frame'], new_quad, canonical_size, smoothing=ctx['args'].smoothing,
                    rebase_after_weak=ctx['args'].rebase_after_weak,
                    rebase_on_recovery=not ctx['args'].no_rebase_on_recovery)
                ctx['fields'].clear()
                ctx['quality_buf'].clear()  # old frames used the previous canonical size
                ctx['dirty'] = True
                refresh_display_params(ctx)
            except RuntimeError as e:
                print(f"Could not use that reference frame: {e}")
    elif action == 'ocr_faster':
        ctx['ocr_every'] = max(1, ctx['ocr_every'] - 1)
    elif action == 'ocr_slower':
        ctx['ocr_every'] += 1
    elif action == 'quit':
        return True
    return False


# --------------------------------------------------------------------------
# Main
# --------------------------------------------------------------------------

def parse_args():
    p = argparse.ArgumentParser(
        description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    p.add_argument("source", help="Path to a video file, or a camera index (e.g. 0)")
    p.add_argument("--ocr", choices=["tesseract", "easyocr"], default="tesseract",
                    help="OCR backend (default: tesseract)")
    p.add_argument("--gpu", action="store_true",
                    help="Use GPU for the easyocr backend (ignored for tesseract)")
    p.add_argument("--output", default="extracted_data.csv",
                    help="File to append logged readings to. A .xlsx path writes an Excel "
                         "workbook (needs openpyxl); any other extension writes CSV "
                         "(default: extracted_data.csv)")
    p.add_argument("--ocr-every", dest="ocr_every", type=int, default=5,
                    help="Run OCR every N frames while playing (default: 5)")
    p.add_argument("--upscale", type=float, default=3.0,
                    help="Upscale factor applied before OCR (default: 3.0)")
    p.add_argument("--no-auto-invert", action="store_true",
                    help="Disable automatic light/dark polarity detection before OCR")
    p.add_argument("--psm", type=int, default=11,
                    help="Tesseract page-segmentation mode (default: 11 = sparse text, "
                         "reads numbers scattered across a multi-value panel). Use 7 for a "
                         "single tightly-cropped value, 6 for a uniform text block. "
                         "(Ignored by the easyocr backend.)")
    p.add_argument("--whitelist", default="0123456789.-:",
                    help="Characters OCR is allowed to output (default: digits plus . - :). "
                         "Pass an empty string (--whitelist \"\") to also capture letters, "
                         "e.g. parameter-name labels like 'U-rms(V)'.")
    p.add_argument("--grid", default=None, metavar="RxC",
                    help="Pre-populate a grid of fields over the display, e.g. --grid 3x4 "
                         "for 3 rows by 4 columns. Each cell becomes its own named column "
                         "(r1c1, r1c2, ...) that you can drag to fit and rename. Best for "
                         "structured multi-value panels; drag any cell corner to adjust it.")
    p.add_argument("--max-width", type=int, default=1280,
                    help="Downscale incoming frames to this width for speed (default: 1280, 0=off)")
    p.add_argument("--smoothing", type=float, default=0.6,
                    help="Jitter-damping factor for the tracked homography, 0-0.97 (default: 0.6)")
    p.add_argument("--sharpness-window", type=int, default=5,
                    help="How many recent frames to pick the sharpest from before OCR, "
                         "when --stack-frames is 1 (default: 5)")
    p.add_argument("--stack-frames", type=int, default=3,
                    help="Median-stack this many recent (already aligned) frames before "
                         "OCR to fight blur/noise that varies frame to frame; set to 1 "
                         "to disable stacking and just pick the sharpest frame instead "
                         "(default: 3)")
    p.add_argument("--rebase-after-weak", type=int, default=20,
                    help="Auto-adopt a fresh reference frame after this many consecutive "
                         "weakly-matched frames, so slow long-term drift doesn't eventually "
                         "stall tracking; 0 disables proactive rebasing (default: 20)")
    p.add_argument("--no-rebase-on-recovery", action="store_true",
                    help="Don't auto-adopt a fresh reference frame right after recovering "
                         "from a tracking loss")
    p.add_argument("--no-auto-log", action="store_true",
                    help="Disable automatic logging. By default a row is written to the "
                         "CSV whenever a field's reading changes, so playing a clip through "
                         "captures every distinct reading with no key presses. With this "
                         "flag, rows are only written when you press 'e'/Log.")
    return p.parse_args()


def main():
    args = parse_args()
    source = int(args.source) if args.source.isdigit() else args.source
    cap = cv2.VideoCapture(source)
    if not cap.isOpened():
        print(f"Could not open video source: {args.source}")
        sys.exit(1)

    ok, frame = cap.read()
    if not ok:
        print("Could not read the first frame.")
        sys.exit(1)
    frame = resize_max_width(frame, args.max_width)

    # Preflight the OCR backend and the output file BEFORE the user picks
    # corners, so a missing backend or a locked/read-only output (e.g. the CSV
    # is open in Excel) fails immediately instead of after all that work.
    try:
        backend = build_ocr_backend(args)
    except RuntimeError as e:
        print(f"OCR backend error: {e}")
        cap.release()
        sys.exit(1)

    try:
        logger = build_row_logger(args.output)
    except (RuntimeError, OSError) as e:
        print(f"Cannot open output file {args.output!r}: {e}")
        if isinstance(e, PermissionError):
            print("It's most likely open in Excel or another program -- close it, "
                  "or pass --output <other.csv> / a .xlsx path.")
        cap.release()
        sys.exit(1)
    # header_written / columns are established lazily on the first logged row
    # (adopting an existing file's header if we're appending to one).
    log_state: dict = {}

    cv2.namedWindow("Original", cv2.WINDOW_NORMAL)
    cv2.namedWindow("Rectified", cv2.WINDOW_NORMAL)

    # Loop so a selection that can't be turned into a usable tracker (too
    # little texture, or a degenerate/collinear quad) re-prompts instead of
    # crashing on confirm -- mirrors the guarded re-pick ('r') path.
    tracker = None
    while tracker is None:
        quad = select_corners_interactively(frame)
        if quad is None:
            print("No quadrilateral selected -- exiting.")
            logger.close()
            cap.release()
            cv2.destroyAllWindows()
            sys.exit(0)
        try:
            canonical_size = canonical_size_from_quad(quad)
            tracker = ReferenceTracker(frame, quad, canonical_size, smoothing=args.smoothing,
                                        rebase_after_weak=args.rebase_after_weak,
                                        rebase_on_recovery=not args.no_rebase_on_recovery)
        except (RuntimeError, cv2.error, MemoryError, ValueError) as e:
            print(f"That selection didn't work ({e}). Please pick the 4 corners again.")
            tracker = None

    initial_fields: List[Field] = []
    if args.grid:
        try:
            grid_rows, grid_cols = parse_grid_spec(args.grid)
        except ValueError as e:
            print(f"Bad --grid value {args.grid!r}: {e}")
            cap.release()
            cv2.destroyAllWindows()
            sys.exit(1)
        initial_fields = build_grid_fields(grid_rows, grid_cols,
                                           tracker.canonical_w, tracker.canonical_h)
        print(f"Created a {grid_rows}x{grid_cols} grid of {len(initial_fields)} fields "
              "-- drag any cell corner to fit it, press 'n' to rename the last one.")

    sharpness_window = max(1, args.sharpness_window)
    stack_frames = max(1, args.stack_frames)

    mouse_state: dict = {}
    scale_ref: List[float] = [1.0]
    image_h_ref: List[int] = [0]
    buttons_ref: List[List[Button]] = [[]]
    fields_ref: List[List[Field]] = [initial_fields]
    canon_ref: List[int] = [tracker.canonical_w, tracker.canonical_h]
    cv2.setMouseCallback("Rectified",
                         make_composite_mouse_callback(mouse_state, scale_ref, image_h_ref,
                                                       buttons_ref, fields_ref, canon_ref))

    ctx = {
        'args': args, 'cap': cap, 'tracker': tracker, 'fields': initial_fields,
        'paused': False, 'frame_idx': 0, 'ocr_every': max(1, args.ocr_every),
        'quality_buf': deque(maxlen=max(sharpness_window, stack_frames, 1)),
        'dirty': True, 'frame': frame, 'logger': logger,
        'log_state': log_state, 'flash_text': None, 'flash_until': 0.0,
        'auto_log': not args.no_auto_log,
        'scale_ref': scale_ref, 'image_h_ref': image_h_ref, 'buttons_ref': buttons_ref,
        'canon_ref': canon_ref,
    }
    refresh_display_params(ctx)

    frame_times: deque = deque(maxlen=30)
    last_ocr_frame = -10**9

    print("Ready. Space = pause/resume, click/drag on 'Rectified' to add a field, "
          "'e' = log, 'c' = clear fields, 'r' = re-pick corners, 'q' = quit -- "
          "or use the on-screen buttons under the Rectified view.")

    while True:
        if not ctx['paused']:
            ok, frame = cap.read()
            if not ok:
                print("End of stream.")
                break
            frame = resize_max_width(frame, args.max_width)
            ctx['frame_idx'] += 1
            frame_times.append(time.time())
        ctx['frame'] = frame

        tracker = ctx['tracker']
        H, quad_in_frame, tracked_ok = tracker.register(frame)
        rectified = cv2.warpPerspective(frame, H, (tracker.canonical_w, tracker.canonical_h))
        cur_sharpness = sharpness_score(rectified)

        if not ctx['paused']:
            ctx['quality_buf'].append((rectified, cur_sharpness))

        should_run_ocr = False
        if not ctx['paused'] and ctx['frame_idx'] - last_ocr_frame >= ctx['ocr_every']:
            should_run_ocr = True
            last_ocr_frame = ctx['frame_idx']
        elif ctx['paused'] and ctx['dirty']:
            should_run_ocr = True

        if should_run_ocr:
            qbuf = ctx['quality_buf']
            if qbuf:
                ocr_source = (stacked_frame(qbuf, stack_frames) if stack_frames > 1
                              else pick_best_frame(qbuf, sharpness_window))
            else:
                ocr_source = rectified  # buffer still empty on the very first frame(s)
            run_ocr_and_update_fields(ocr_source, ctx['fields'], backend, args)
            ctx['dirty'] = False
            if ctx['auto_log']:
                maybe_auto_log(ctx)

        # ---- "Original" window: tracked outline + stats HUD + flash banner ----
        display_frame = frame.copy()
        draw_original_overlay(display_frame, quad_in_frame, tracked_ok, ctx['paused'], tracker.rebase_count)

        if len(frame_times) >= 2 and frame_times[-1] > frame_times[0]:
            fps = (len(frame_times) - 1) / (frame_times[-1] - frame_times[0])
            fps_line = f"FPS: {fps:4.1f}"
        else:
            fps_line = "FPS: --  (paused)" if ctx['paused'] else "FPS: --"
        ocr_mode = f"stack x{stack_frames}" if stack_frames > 1 else f"best-of-{sharpness_window}"
        stats_lines = [
            fps_line,
            f"Matches: {tracker.last_good_matches}  Inliers: {tracker.last_inlier_count}",
            f"Weak streak: {tracker.consec_weak}  Rebases: {tracker.rebase_count}",
            f"Sharpness: {cur_sharpness:6.0f}",
            f"OCR every {ctx['ocr_every']}f | {ocr_mode}",
            f"Fields: {len(ctx['fields'])}   Frame: {ctx['frame_idx']}",
            f"Rows logged: {ctx['log_state'].get('rows_logged', 0)}",
        ]
        draw_stats_panel(display_frame, stats_lines)
        if ctx['flash_text'] and time.time() < ctx['flash_until']:
            draw_flash_banner(display_frame, ctx['flash_text'])
        else:
            ctx['flash_text'] = None
        cv2.imshow("Original", display_frame)

        # ---- "Rectified" window: annotated crop, scaled up, + toolbar ----
        rectified_annotated = rectified.copy()
        draw_rectified_overlay(rectified_annotated, ctx['fields'], mouse_state)
        disp_w, disp_h = ctx['disp_w'], ctx['disp_h']
        interp = cv2.INTER_NEAREST if scale_ref[0] > 1.0 else cv2.INTER_AREA
        rectified_disp = cv2.resize(rectified_annotated, (disp_w, disp_h), interpolation=interp)

        pressed_cmd = None
        if mouse_state.get("flash_cmd") and time.time() < mouse_state.get("flash_button_until", 0):
            pressed_cmd = mouse_state["flash_cmd"]
        toolbar_img = draw_toolbar(disp_w, buttons_ref[0], ctx['paused'], pressed_cmd)
        cv2.imshow("Rectified", np.vstack([rectified_disp, toolbar_img]))

        # ---- input: keyboard ----
        key = cv2.waitKey(1 if not ctx['paused'] else 30) & 0xFF
        if key == ord('n') and ctx['fields']:
            new_name = input(f"Rename last field (was '{ctx['fields'][-1].name}'): ").strip()
            if new_name:
                ctx['fields'][-1].name = new_name
        else:
            action = COMMAND_KEYS.get(key)
            if action and apply_action(action, ctx):
                break

        # ---- input: toolbar button click ----
        btn_cmd = mouse_state.pop("button_cmd", None)
        if btn_cmd:
            btn_action = BUTTON_TO_ACTION.get(btn_cmd)
            if btn_action and apply_action(btn_action, ctx):
                break

        # ---- a completed drag/click defines (or resizes) a field ----
        if mouse_state.get("completed_rect"):
            x0, y0, x1, y1 = mouse_state.pop("completed_rect")
            if abs(x1 - x0) < 5 and abs(y1 - y0) < 5:
                dw, dh = 80, 40  # plain click -> default-size box centered there
                x, y = max(0, int(x0 - dw // 2)), max(0, int(y0 - dh // 2))
                w, h = dw, dh
            else:
                x, y = int(min(x0, x1)), int(min(y0, y1))
                w, h = int(abs(x1 - x0)), int(abs(y1 - y0))
            if w > 4 and h > 4:
                ctx['fields'].append(Field(name=f"field{len(ctx['fields']) + 1}", rect=(x, y, w, h)))
                ctx['dirty'] = True

    cap.release()
    logger.close()
    cv2.destroyAllWindows()
    print(f"Data saved to {Path(args.output).resolve()}")


if __name__ == "__main__":
    main()
